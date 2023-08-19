# 엑셀을 사용하기 위한 모듈
from openpyxl import load_workbook


class read_excel():
    def __init__(self):
        pass

    def run(self):
        list = []
        wb = load_workbook(filename="data_j.xlsx")
        ws = wb[wb.sheetnames[0]]
        idx = 1
        while True:
            idx = idx+1
            code = ws[f'B{idx}'].value
            if code == None or str(code).strip() == '':
                break
            #list.append({
            #    "code": code,
            #    "date": str(ws[f'A{idx}'].value),
            #    "name": str(ws[f'C{idx}'].value),
            #    "stockcode": str(ws[f'D{idx}'].value),
            #    "type33code": str(ws[f'E{idx}'].value),
            #    "type17code": str(ws[f'F{idx}'].value),
            #    "typescalecode": str(ws[f'G{idx}'].value)
            #})
            list.append(tuple([
                code, str(ws[f'A{idx}'].value), str(ws[f'C{idx}'].value), str(ws[f'D{idx}'].value), str(ws[f'E{idx}'].value), str(ws[f'F{idx}'].value), str(ws[f'G{idx}'].value)
            ]))
            pass
        return list
